"""Carrinho, pedidos, status, histórico e exportação."""

from io import BytesIO

from flask import (
    Blueprint,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for
)

from openpyxl import Workbook
from openpyxl.styles import Font

from models import pedido as pedido_model
from models import prato as prato_model


# Blueprint dos pedidos.
pedido_bp = Blueprint(
    "pedido",
    __name__,
    url_prefix="/pedidos"
)


def _carrinho():

    # Obtém o carrinho da sessão.
    # Caso ainda não exista, cria uma lista vazia.
    return session.setdefault(
        "carrinho",
        []
    )


@pedido_bp.get("")
def listar():

    # Lista os pedidos registrados.
    pedidos = pedido_model.listar()

    return render_template(
        "pedidos.html",
        pedidos=pedidos
    )


@pedido_bp.get("/novo")
def novo():

    # Obtém o carrinho.
    carrinho = _carrinho()

    # Calcula o total.
    total = sum(
        item["preco_centavos"]
        * item["quantidade"]
        for item in carrinho
    )

    # Exibe os pratos e o carrinho.
    return render_template(
        "novo_pedido.html",
        pratos=prato_model.listar(),
        carrinho=carrinho,
        total=total
    )


@pedido_bp.post(
    "/carrinho/adicionar"
)
def adicionar():

    try:

        # Converte o ID para inteiro.
        prato_id = int(
            request.form.get(
                "prato_id",
                ""
            )
        )

        # Converte a quantidade.
        quantidade = int(
            request.form.get(
                "quantidade",
                ""
            )
        )

    except ValueError:

        prato_id = 0
        quantidade = 0

    # Busca o prato.
    prato = prato_model.buscar(
        prato_id
    )

    # Verifica os dados.
    if prato is None or quantidade <= 0:

        flash(
            "Selecione um prato e uma quantidade válida.",
            "erro"
        )

        return redirect(
            url_for("pedido.novo")
        )

    # Obtém o carrinho.
    carrinho = _carrinho()

    # Procura se o prato já está no carrinho.
    existente = next(
        (
            item
            for item in carrinho
            if item["prato_id"] == prato_id
        ),
        None
    )

    # Se já existe, soma a quantidade.
    nova_quantidade = quantidade

    if existente:

        nova_quantidade += (
            existente["quantidade"]
        )

    # Verifica o limite.
    # Neste ponto não existe estoque de prato:
    # o controle será realizado pelos ingredientes
    # no momento da finalização.
    if existente:

        existente["quantidade"] = nova_quantidade

    else:

        carrinho.append({
            "prato_id": prato["id"],
            "nome": prato["nome"],
            "preco_centavos": prato["preco_centavos"],
            "quantidade": quantidade
        })

    # Atualiza a sessão.
    session["carrinho"] = carrinho

    flash(
        "Prato adicionado ao pedido.",
        "sucesso"
    )

    return redirect(
        url_for("pedido.novo")
    )


@pedido_bp.post(
    "/carrinho/<int:indice>/remover"
)
def remover(indice):

    # Obtém o carrinho.
    carrinho = _carrinho()

    # Verifica se o índice existe.
    if (
        indice < 0
        or indice >= len(carrinho)
    ):
        abort(404)

    # Remove o item.
    carrinho.pop(indice)

    # Atualiza a sessão.
    session["carrinho"] = carrinho

    return redirect(
        url_for("pedido.novo")
    )


@pedido_bp.post(
    "/carrinho/limpar"
)
def limpar():

    # Remove o carrinho.
    session.pop(
        "carrinho",
        None
    )

    flash(
        "Pedido esvaziado.",
        "sucesso"
    )

    return redirect(
        url_for("pedido.novo")
    )


@pedido_bp.post("/finalizar")
def finalizar():

    # Obtém o carrinho.
    carrinho = _carrinho()

    # Não permite pedido vazio.
    if not carrinho:

        flash(
            "O pedido está vazio.",
            "erro"
        )

        return redirect(
            url_for("pedido.novo")
        )

    try:

        # Cria o pedido no banco.
        pedido_id = pedido_model.finalizar(
            session["funcionario_id"],
            carrinho
        )

    except ValueError as erro:

        # Exibe problemas como estoque insuficiente.
        flash(
            str(erro),
            "erro"
        )

        return redirect(
            url_for("pedido.novo")
        )

    # Limpa o carrinho.
    session.pop(
        "carrinho",
        None
    )

    flash(
        "Pedido registrado com sucesso.",
        "sucesso"
    )

    return redirect(
        url_for(
            "pedido.detalhes",
            pedido_id=pedido_id
        )
    )


@pedido_bp.get(
    "/<int:pedido_id>"
)
def detalhes(pedido_id):

    # Busca o pedido.
    pedido = pedido_model.buscar(
        pedido_id
    )

    if pedido is None:
        abort(404)

    # Busca os itens.
    itens = pedido_model.listar_itens(
        pedido_id
    )

    return render_template(
        "detalhes_pedido.html",
        pedido=pedido,
        itens=itens
    )


@pedido_bp.post(
    "/<int:pedido_id>/status"
)
def alterar_status(pedido_id):

    # Estados permitidos.
    status_permitidos = {
        "espera",
        "preparado",
        "entregue"
    }

    # Obtém o novo status.
    status = request.form.get(
        "status",
        ""
    )

    # Valida o status.
    if status not in status_permitidos:

        flash(
            "Status inválido.",
            "erro"
        )

        return redirect(
            url_for(
                "pedido.detalhes",
                pedido_id=pedido_id
            )
        )

    # Verifica se o pedido existe.
    if pedido_model.buscar(
        pedido_id
    ) is None:
        abort(404)

    # Atualiza o status.
    pedido_model.atualizar_status(
        pedido_id,
        status
    )

    flash(
        "Status do pedido atualizado.",
        "sucesso"
    )

    return redirect(
        url_for(
            "pedido.detalhes",
            pedido_id=pedido_id
        )
    )


@pedido_bp.get("/exportar")
def exportar():

    # Cria uma planilha.
    wb = Workbook()

    # Obtém a planilha ativa.
    ws = wb.active

    ws.title = "Pedidos"

    # Cabeçalho.
    ws.append([
        "Código",
        "Data e hora",
        "Funcionário",
        "Status",
        "Total (R$)"
    ])

    # Negrito no cabeçalho.
    for cell in ws[1]:
        cell.font = Font(
            bold=True
        )

    # Adiciona os pedidos.
    for p in pedido_model.listar():

        ws.append([
            p["id"],
            p["criado_em"],
            p["funcionario_nome"],
            p["status"],
            p["total_centavos"] / 100
        ])

    # Ajusta larguras.
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15

    # Formata os valores monetários.
    for cell in ws["E"][1:]:
        cell.number_format = "R$ #,##0.00"

    # Cria o arquivo em memória.
    arquivo = BytesIO()

    wb.save(arquivo)

    arquivo.seek(0)

    # Envia o Excel.
    return send_file(
        arquivo,
        as_attachment=True,
        download_name="pedidos.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )