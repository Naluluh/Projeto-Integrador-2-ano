"""CRUD e exportação de pratos."""

from io import BytesIO

from flask import (
    Blueprint,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for
)

from openpyxl import Workbook
from openpyxl.styles import Font

from models import prato as prato_model
from utils import moeda_para_centavos


# Blueprint responsável pelos pratos.
prato_bp = Blueprint(
    "prato",
    __name__,
    url_prefix="/pratos"
)


def _dados_formulario():
    # Obtém o nome do prato.
    nome = request.form.get(
        "nome",
        ""
    ).strip()

    # Obtém a descrição.
    descricao = request.form.get(
        "descricao",
        ""
    ).strip()

    # Obtém a categoria.
    categoria = request.form.get(
        "categoria",
        ""
    ).strip()

    # Obtém o modo de preparo.
    modo_preparo = request.form.get(
        "modo_preparo",
        ""
    ).strip()

    # Converte o preço para centavos.
    preco = moeda_para_centavos(
        request.form.get("preco", "")
    )

    # Verifica se os dados obrigatórios são válidos.
    if (
        not nome
        or not descricao
        or not categoria
        or not modo_preparo
        or preco < 0
    ):
        raise ValueError(
            "Preencha todos os campos com valores válidos."
        )

    return (
        nome,
        descricao,
        categoria,
        preco,
        modo_preparo
    )


@prato_bp.get("")
def listar():

    # Obtém o termo digitado na pesquisa.
    busca = request.args.get(
        "busca",
        ""
    )

    # Busca os pratos no Model.
    pratos = prato_model.listar(
        busca
    )

    # Exibe a página.
    return render_template(
        "pratos.html",
        pratos=pratos,
        busca=busca
    )


@prato_bp.route(
    "/novo",
    methods=("GET", "POST")
)
def novo():

    # Processa o formulário quando enviado.
    if request.method == "POST":

        try:

            # Obtém os dados validados.
            dados = _dados_formulario()

            # Cria o prato.
            prato_model.criar(*dados)

            flash(
                "Prato cadastrado.",
                "sucesso"
            )

            return redirect(
                url_for("prato.listar")
            )

        except (ValueError, TypeError):

            flash(
                "Preencha os dados do prato corretamente.",
                "erro"
            )

    return render_template(
        "prato_form.html",
        prato=None
    )


@prato_bp.route(
    "/<int:prato_id>/editar",
    methods=("GET", "POST")
)
def editar(prato_id):

    # Procura o prato.
    prato = prato_model.buscar(
        prato_id
    )

    # Se não existir, retorna 404.
    if prato is None:
        abort(404)

    if request.method == "POST":

        try:

            # Obtém os novos dados.
            dados = _dados_formulario()

            # Atualiza o prato.
            prato_model.atualizar(
                prato_id,
                *dados
            )

            flash(
                "Prato atualizado.",
                "sucesso"
            )

            return redirect(
                url_for("prato.listar")
            )

        except (ValueError, TypeError):

            flash(
                "Preencha os dados do prato corretamente.",
                "erro"
            )

    return render_template(
        "prato_form.html",
        prato=prato
    )


@prato_bp.post(
    "/<int:prato_id>/excluir"
)
def excluir(prato_id):

    # Verifica se o prato existe.
    if prato_model.buscar(prato_id) is None:
        abort(404)

    # Exclui o prato.
    prato_model.excluir(
        prato_id
    )

    flash(
        "Prato excluído.",
        "sucesso"
    )

    return redirect(
        url_for("prato.listar")
    )


@prato_bp.get("/exportar")
def exportar():

    # Cria uma planilha Excel.
    wb = Workbook()

    # Obtém a planilha ativa.
    ws = wb.active

    # Define seu nome.
    ws.title = "Pratos"

    # Cria o cabeçalho.
    ws.append([
        "Código",
        "Nome",
        "Descrição",
        "Categoria",
        "Preço (R$)",
        "Modo de preparo"
    ])

    # Deixa o cabeçalho em negrito.
    for cell in ws[1]:
        cell.font = Font(
            bold=True
        )

    # Adiciona os pratos.
    for p in prato_model.listar():

        ws.append([
            p["id"],
            p["nome"],
            p["descricao"],
            p["categoria"],
            p["preco_centavos"] / 100,
            p["modo_preparo"]
        ])

    # Ajusta algumas larguras.
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 45

    # Formata os preços.
    for cell in ws["E"][1:]:
        cell.number_format = "R$ #,##0.00"

    # Cria um arquivo em memória.
    arquivo = BytesIO()

    # Salva a planilha nesse arquivo.
    wb.save(arquivo)

    # Volta para o início do arquivo.
    arquivo.seek(0)

    # Envia o Excel para o navegador.
    return send_file(
        arquivo,
        as_attachment=True,
        download_name="pratos.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )